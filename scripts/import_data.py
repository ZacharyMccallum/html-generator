# Load xlsx workbook
import openpyxl

def readData():
    # Create workbook var for import spreadsheet
    workbook = openpyxl.load_workbook(r'./src/user_details.xlsx')
    data = workbook.active




    # Create empty dictionary for adding to
    xlsx_content = {}
    for row in data.iter_rows(min_row=1, min_col=1, max_row=36, max_col=6):
        for cell in row:
            # if row = 1 then save as key, else - save it as a value within the key
            if row != 1:
                xlsx_content[row][cell] = str(cell.value)
            else:
                xlsx_content[row] = cell.value
            print(cell.value, end=" ")
            print("\n")
    print(xlsx_content)

#    for rowOfCellObjects in data:
#        print(rowOfCellObjects)
# #       xlsx_content[str(rowOfCellObjects)] = str(rowOfCellObjects.value)
#        for cellObj in rowOfCellObjects:
#            # Create new dictionary entry, key = row?, val = value?
#            xlsx_content[str(cellObj.row)] = str(cellObj.value)
#            print(rowOfCellObjects[1], cellObj.value)
#    print(xlsx_content)
#    print('---- End of Row ----')
        
#        for col in range(1, data.max_column + 1):
#            print(col)
#            if row != 1:                
#                cell_obj = data.cell(row=row, column=col)
#                print(cell_obj.value, end=" ")
#                # Create values using j along each col
#                xlsx_content[row] = cell_obj
#                return xlsx_content
#            else:
#                # find keys from the first row (column headers)
#                cell_obj = data.cell(row=1, column=col)
#                print(cell_obj.value, end=" ")
#                print(row)
#                xlsx_content[row][col] = cell_obj
#
#                print("Another condition")
readData()