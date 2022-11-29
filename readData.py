import openpyxl
from openpyxl import load_workbook



workbook = openpyxl.load_workbook("user_details.xlsx")

data = workbook.active


for i in range(1, data.max_row + 1):
    print("\n")
    print("Row ", i, " data :")

    for j in range(1, data.max_column + 1):
        cell_obj = data.cell(row=i, column=j)
        print(cell_obj.value, end=" ")